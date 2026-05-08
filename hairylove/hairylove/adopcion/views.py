import re
from urllib import request
from urllib.parse import urljoin
from urllib.error import URLError, HTTPError
from html import unescape
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db import models
from rest_framework import viewsets, permissions, status, pagination
from rest_framework.decorators import action
from rest_framework.response import Response
import openpyxl
from openpyxl.utils.datetime import from_excel
from datetime import datetime
from .models import Mascota, Adopcion, Calificacion, ChatMessage
from .serializers import MascotaSerializer, AdopcionSerializer
from .forms import MascotaAdopcionForm, AdopcionForm, CalificacionForm, ChatMessageForm, CargaMasivaForm
from .razas import RAZAS_POR_ESPECIE, ESPECIES
from usuarios.models import Usuario, Criador
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from django.contrib.staticfiles.finders import find
from datetime import datetime, date, timedelta
from openpyxl.utils.datetime import from_excel
from io import BytesIO
from django.conf import settings
import os
import uuid
from django.core.files.base import ContentFile


def parse_excel_date(value, workbook=None):
    """Parsea una fecha proveniente de Excel en un objeto date."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value, workbook.epoch if workbook else datetime(1899, 12, 30)).date()
        except Exception:
            pass
    if isinstance(value, str):
        text = value.strip()
        if text == '':
            return None
        # If it contains time, keep only the date part.
        text = text.split()[0]
        formatos = [
            '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
            '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y',
            '%m/%d/%Y', '%m-%d-%Y', '%m.%d.%Y'
        ]
        for fmt in formatos:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        # Try ISO parser fallback for formats like 2020-05-15T00:00:00
        try:
            return datetime.fromisoformat(text).date()
        except Exception:
            pass
    raise ValueError(f'No se pudo parsear la fecha: {value}')

def parse_boolean(value):
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    return text in ['si', 'sí', 'yes', 'true', '1', 'verdadero']



def parse_boolean(value):
    """Convierte un valor a booleano (Si/No, Yes/No, True/False, 1/0)"""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    return text in ['si', 'sí', 'yes', 'true', '1', 'verdadero']


def extract_image_url_from_html(page_url, html_text):
    html_text = html_text.replace('\n', ' ')
    patterns = [
        r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+name=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+name=["\']twitter:image["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+property=["\']twitter:image["\'][^>]+content=["\']([^"\']+)["\']',
        r'<img[^>]+data-src=["\']([^"\']+)["\']',
        r'<img[^>]+src=["\']([^"\']+)["\']',
    ]
    for pattern in patterns:
        match = re.search(pattern, html_text, flags=re.I)
        if match:
            url = unescape(match.group(1).strip())
            if url:
                return urljoin(page_url, url)
    return None


def download_image_from_url(image_url):
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    req = request.Request(image_url, headers=headers)
    response = request.urlopen(req, timeout=15)
    content = response.read()
    content_type = response.getheader('Content-Type', '') or ''

    if content_type.lower().startswith('image/'):
        return content, content_type

    html_text = content.decode('utf-8', errors='ignore')
    extracted = extract_image_url_from_html(image_url, html_text)
    if extracted:
        req2 = request.Request(extracted, headers={**headers, 'Referer': image_url})
        response2 = request.urlopen(req2, timeout=15)
        content2 = response2.read()
        content_type2 = response2.getheader('Content-Type', '') or ''
        if content_type2.lower().startswith('image/'):
            return content2, content_type2
    raise ValueError('No se encontró una imagen en la URL proporcionada')


def validate_and_map_estado_salud(valor):
    """
    Valida y mapea el valor de Estado_Salud a uno de los valores permitidos.
    Si no es válido, retorna el valor por defecto 'Buena'.
    """
    if not valor:
        return 'Buena'
    
    texto = str(valor).strip().lower()
    
    # Mapeo de posibles valores en el Excel
    mapeo = {
        'excelente': 'Excelente',
        'buena': 'Buena',
        'regular': 'Regular',
        'muy buena': 'Excelente',
        'muy buena salud': 'Excelente',
        'buena salud': 'Buena',
        'regular salud': 'Regular',
        'mala': 'Regular',
        'pobre': 'Regular',
        'ok': 'Buena',
        'bien': 'Buena',
        'malo': 'Regular',
    }
    
    # Buscar coincidencia exacta o parcial
    for clave, val_mapeo in mapeo.items():
        if clave in texto or texto in clave:
            return val_mapeo
    
    # Si no encuentra coincidencia, retorna el valor por defecto
    return 'Buena'


def validate_and_map_genero(valor):
    """Valida y mapea el género a uno de los valores permitidos."""
    if not valor:
        return 'Macho'
    
    texto = str(valor).strip().lower()
    
    if texto in ['macho', 'm', 'male', 'masculino']:
        return 'Macho'
    elif texto in ['hembra', 'h', 'female', 'femenino', 'f']:
        return 'Hembra'
    
    return 'Macho'


def validate_and_map_tamaño(valor):
    """Valida y mapea el tamaño a uno de los valores permitidos."""
    if not valor:
        return 'Mediano'
    
    texto = str(valor).strip().lower()
    
    if 'pequeño' in texto or 'pequeño' in texto or 'small' in texto or 'chico' in texto:
        return 'Pequeño'
    elif 'grande' in texto or 'big' in texto or 'l' == texto or 'lg' in texto:
        return 'Grande'
    else:
        return 'Mediano'


def validate_and_map_origen(valor):
    """Valida y mapea el origen a uno de los valores permitidos."""
    if not valor:
        return 'Criador'
    
    texto = str(valor).strip().lower()
    opciones = ['criador', 'refugio', 'rescate', 'abandono']
    
    for opcion in opciones:
        if opcion in texto or texto in opcion:
            return opcion.capitalize()
    
    return 'Criador'


from openpyxl import Workbook
import random


# ==================== VISTAS TRADICIONALES ====================

def mascotas(request):
    mascotas = Mascota.objects.all()
    context = {
        'mascotas': mascotas
    }
    return render(request, 'usuarios/mascotas.html', context)


def formulario_adopcion(request):
    """Redirige a la página de mascotas disponibles para adopción"""
    return redirect('mascotas_adopcion_disponibles')


@login_required
def registrar_mascota_adopcion(request):
    """Vista para que criadores registren mascotas en adopción"""
    # Verificar que el usuario sea Criador
    from usuarios.models import Criador
    
    usuario = request.user
    try:
        criador = Criador.objects.get(user=usuario)
    except:
        messages.error(request, "Solo los criadores pueden registrar mascotas en adopción")
        return redirect('index')
    
    if request.method == 'POST':
        form = MascotaAdopcionForm(request.POST, request.FILES)
        if form.is_valid():
            mascota = form.save(commit=False)
            mascota.idCriador = criador.idCriador
            mascota.disponible = True
            mascota.fecha_creacion = timezone.now()
            mascota.save()
            
            messages.success(request, f"{mascota.Nombre_Mascota} ha sido registrada exitosamente en adopcion.")
            return redirect('mis_mascotas_adopcion')
        else:
            messages.error(request, "Error al registrar la mascota. Por favor verifica los datos.")
    else:
        form = MascotaAdopcionForm()
    
    context = {'form': form}
    return render(request, 'adopcion/registrar_mascota.html', context)


@login_required
def mis_mascotas_adopcion(request):
    """Vista para que criadores vean sus mascotas en adopción"""
    from usuarios.models import Criador
    
    usuario = request.user
    try:
        criador = Criador.objects.get(user=usuario)
    except:
        messages.error(request, "Solo los criadores pueden acceder a esta función")
        return redirect('index')
    
    mascotas = Mascota.objects.filter(idCriador=criador.idCriador, disponible=True)
    context = {'mascotas': mascotas}
    return render(request, 'adopcion/mis_mascotas_adopcion.html', context)


def mascotas_adopcion_disponibles(request):
    """Vista para mostrar mascotas disponibles en adopción"""
    # Filtrar mascotas disponibles que NO han sido adoptadas
    mascotas = Mascota.objects.filter(
        disponible=True
    ).exclude(
        adopcion__Estado='Aprobada'
    ).distinct()
    
    # Filtros opcionales
    especie = request.GET.get('especie')
    tamaño = request.GET.get('tamaño')
    genero = request.GET.get('genero')
    puntuacion_minima = request.GET.get('puntuacion_minima')
    
    if especie:
        mascotas = mascotas.filter(Especie__icontains=especie)
    if tamaño:
        mascotas = mascotas.filter(Tamaño__icontains=tamaño)
    if genero:
        mascotas = mascotas.filter(Genero=genero)
    
    # Filtro por calificación del criador
    if puntuacion_minima:
        try:
            puntuacion_minima = float(puntuacion_minima)
            # Filtrar mascotas cuyos criadores tienen puntuación >= a la mínima
            mascotas = mascotas.filter(
                idCriador__in=[c.idCriador for c in Criador.objects.filter(
                    user__puntuacion_promedio__gte=puntuacion_minima
                )]
            )
        except (ValueError, TypeError):
            pass
    
    # Ordenar por más recientes
    mascotas = mascotas.order_by('-fecha_creacion')

    # Limpiar nombres repetidos y eliminar sufijos numéricos usados en datos de prueba.
    nombre_grupos = {}
    for mascota in mascotas:
        nombre_limpio = re.sub(r'\s+\d+$', '', mascota.Nombre_Mascota).strip()
        nombre_grupos.setdefault(nombre_limpio, []).append(mascota)

    for nombre_limpio, grupo in nombre_grupos.items():
        if len(grupo) == 1:
            grupo[0].display_name = nombre_limpio
        else:
            nombres_usados = set()
            for index, mascota in enumerate(grupo, start=1):
                # Asignar un nombre base con raza y desambiguar si ya existe.
                display_base = f"{nombre_limpio} ({mascota.Raza})"
                if display_base in nombres_usados:
                    mascota.display_name = f"{nombre_limpio} ({mascota.Raza} {index})"
                else:
                    mascota.display_name = display_base
                    nombres_usados.add(display_base)

    context = {
        'mascotas': mascotas,
        'total_mascotas': mascotas.count(),
        'puntuacion_minima': puntuacion_minima or '',
        'especie': especie or '',
        'tamaño': tamaño or '',
        'genero': genero or '',
    }
    return render(request, 'adopcion/mascotas_disponibles.html', context)


def detalles_mascota(request, mascota_id):
    """Vista para ver detalles de una mascota disponible"""
    try:
        mascota = Mascota.objects.get(idMascota=mascota_id, disponible=True)
    except Mascota.DoesNotExist:
        messages.error(request, "Mascota no encontrada")
        return redirect('mascotas_adopcion_disponibles')
    
    # Verificar si el usuario autenticado ya tiene una solicitud pendiente
    solicitud_existente = None
    es_criador_propietario = False
    
    if request.user.is_authenticated:
        # Verificar si es el criador propietario
        if hasattr(request.user, 'idUsuario') and request.user.idUsuario == mascota.idCriador:
            es_criador_propietario = True
        else:
            solicitud_existente = Adopcion.objects.filter(
                idMascota=mascota,
                idPropietario=request.user.idUsuario,
                Estado__in=['Pendiente', 'Aprobada']
            ).first()
    
    context = {
        'mascota': mascota,
        'solicitud_existente': solicitud_existente,
        'es_criador_propietario': es_criador_propietario
    }
    return render(request, 'adopcion/detalles_mascota.html', context)


@login_required
def solicitar_adopcion(request, mascota_id):
    """Vista para procesar una solicitud de adopción"""
    try:
        mascota = Mascota.objects.get(idMascota=mascota_id, disponible=True)
    except Mascota.DoesNotExist:
        messages.error(request, "Mascota no encontrada")
        return redirect('mascotas_adopcion_disponibles')
    
    # Verificar si ya existe una solicitud pendiente
    adopcion_existente = Adopcion.objects.filter(
        idMascota=mascota,
        idPropietario=request.user.idUsuario,
        Estado__in=['Pendiente', 'Confirmada']
    ).first()
    
    if adopcion_existente:
        messages.warning(request, "Ya tienes una solicitud pendiente para esta mascota")
        return redirect('detalles_mascota', mascota_id=mascota_id)
    
    if request.method == 'POST':
        form = AdopcionForm(request.POST)
        if form.is_valid():
            adopcion = form.save(commit=False)
            adopcion.idMascota = mascota
            adopcion.idPropietario = request.user.idUsuario
            adopcion.idCriador = mascota.idCriador
            adopcion.Estado = 'Pendiente'
            adopcion.Estado_Solicitud = 'En revisión'
            adopcion.Fecha_Solicitud = date.today()
            adopcion.Fecha_Adopción = date.today() + timedelta(days=2)
            adopcion.Fecha_Entrega = date.today() + timedelta(days=7)
            adopcion.Fuente_Mascota = mascota.Origen
            adopcion.save()
            
            messages.success(request, "Tu solicitud de adopcion ha sido registrada. El criador la revisara pronto.")
            return redirect('mis_adopciones')
        else:
            messages.error(request, "Error al procesar la solicitud de adopción")
    else:
        form = AdopcionForm()
    
    context = {
        'form': form,
        'mascota': mascota
    }
    return render(request, 'adopcion/solicitar_adopcion.html', context)


@login_required
def mis_adopciones(request):
    """Vista para ver las adopciones del usuario actual"""
    adopciones = Adopcion.objects.filter(idPropietario=request.user.idUsuario).order_by('-Fecha_Solicitud')
    
    # Agregar información de calificación a cada adopción
    for adopcion in adopciones:
        calificacion = Calificacion.objects.filter(
            usuario_califica=request.user,
            adopcion=adopcion
        ).first()
        adopcion.calificacion_del_usuario = calificacion
    
    for adopcion in adopciones:
        actualizar_estado_entrega(adopcion)

    context = {
        'adopciones': adopciones,
        'total_adopciones': adopciones.count()
    }
    return render(request, 'adopcion/mis_adopciones.html', context)


@login_required
def chat_lista(request):
    """Lista de conversaciones disponibles para el usuario."""
    from usuarios.models import Usuario

    usuario_actual = request.user
    contactos = Usuario.objects.none()

    if usuario_actual.tipo == 'Propietario':
        criador_ids = Adopcion.objects.filter(idPropietario=usuario_actual.idUsuario).values_list('idCriador', flat=True).distinct()
        contactos = Usuario.objects.filter(idUsuario__in=[cid for cid in criador_ids if cid]).exclude(idUsuario=usuario_actual.idUsuario)
    elif usuario_actual.tipo == 'Criador':
        propietario_ids = Adopcion.objects.filter(idCriador=usuario_actual.idUsuario).values_list('idPropietario', flat=True).distinct()
        contactos = Usuario.objects.filter(idUsuario__in=[pid for pid in propietario_ids if pid]).exclude(idUsuario=usuario_actual.idUsuario)
    else:
        contactos = Usuario.objects.exclude(idUsuario=usuario_actual.idUsuario)[:20]

    context = {
        'usuarios_chat': contactos,
    }
    return render(request, 'adopcion/chat_lista.html', context)


@login_required
def chat_mensajes(request, usuario_id):
    """Vista de chat entre propietario y criador"""
    from usuarios.models import Usuario

    usuario_actual = request.user
    try:
        usuario_destino = Usuario.objects.get(idUsuario=usuario_id)
    except Usuario.DoesNotExist:
        messages.error(request, 'Usuario no encontrado')
        return redirect('index')

    # Verificar que uno sea propietario y otro criador
    if usuario_actual.tipo not in ['Propietario', 'Criador'] or usuario_destino.tipo not in ['Propietario', 'Criador']:
        messages.error(request, 'Chat no disponible entre estos usuarios')
        return redirect('index')

    mensajes = ChatMessage.objects.filter(
        (models.Q(remitente=usuario_actual) & models.Q(receptor=usuario_destino)) |
        (models.Q(remitente=usuario_destino) & models.Q(receptor=usuario_actual))
    ).order_by('timestamp')

    # Marcar como leídos
    ChatMessage.objects.filter(receptor=usuario_actual, remitente=usuario_destino, leido=False).update(leido=True)

    # Lista de contactos con los que el usuario puede chatear
    usuarios_chat = Usuario.objects.none()
    if usuario_actual.tipo == 'Propietario':
        criador_ids = Adopcion.objects.filter(idPropietario=usuario_actual.idUsuario).values_list('idCriador', flat=True).distinct()
        usuarios_chat = Usuario.objects.filter(idUsuario__in=[cid for cid in criador_ids if cid]).exclude(idUsuario=usuario_actual.idUsuario)
    elif usuario_actual.tipo == 'Criador':
        propietario_ids = Adopcion.objects.filter(idCriador=usuario_actual.idUsuario).values_list('idPropietario', flat=True).distinct()
        usuarios_chat = Usuario.objects.filter(idUsuario__in=[pid for pid in propietario_ids if pid]).exclude(idUsuario=usuario_actual.idUsuario)
    else:
        usuarios_chat = Usuario.objects.exclude(idUsuario=usuario_actual.idUsuario)[:20]

    if request.method == 'POST':
        form = ChatMessageForm(request.POST)
        if form.is_valid():
            texto = form.cleaned_data['mensaje']
            ChatMessage.objects.create(
                remitente=usuario_actual,
                receptor=usuario_destino,
                mensaje=texto
            )
            return redirect('chat_mensajes', usuario_id=usuario_id)
    else:
        form = ChatMessageForm()

    context = {
        'usuario_destino': usuario_destino,
        'mensajes': mensajes,
        'form': form,
        'usuarios_chat': usuarios_chat,
    }
    return render(request, 'adopcion/chat_mensajes.html', context)


def actualizar_estado_entrega(adopcion):
    """Actualiza el estado de entrega si el tiempo de entrega ya llegó."""
    if adopcion.Estado == 'Aprobada' and adopcion.Estado_Solicitud == 'En camino' and adopcion.Fecha_Entrega <= date.today():
        adopcion.Estado_Solicitud = 'Completada'
        adopcion.save()


@login_required
def chat_mensajes_api(request, usuario_id):
    """Devuelve los mensajes del chat en JSON para actualización automática."""
    try:
        usuario_destino = Usuario.objects.get(idUsuario=usuario_id)
    except Usuario.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)

    usuario_actual = request.user
    if usuario_actual.tipo not in ['Propietario', 'Criador'] or usuario_destino.tipo not in ['Propietario', 'Criador']:
        return JsonResponse({'error': 'Chat no disponible entre estos usuarios'}, status=403)

    last_id = request.GET.get('last_id')
    mensajes_query = ChatMessage.objects.filter(
        (models.Q(remitente=usuario_actual) & models.Q(receptor=usuario_destino)) |
        (models.Q(remitente=usuario_destino) & models.Q(receptor=usuario_actual))
    )

    if last_id and last_id.isdigit():
        mensajes_query = mensajes_query.filter(idChat__gt=int(last_id))

    mensajes = mensajes_query.order_by('timestamp')
    ChatMessage.objects.filter(receptor=usuario_actual, remitente=usuario_destino, leido=False).update(leido=True)

    mensajes_data = [
        {
            'id': mensaje.idChat,
            'sender_id': mensaje.remitente.idUsuario,
            'sender_name': mensaje.remitente.nombre,
            'message': mensaje.mensaje,
            'timestamp': mensaje.timestamp.strftime('%d/%m/%Y %H:%M'),
        }
        for mensaje in mensajes
    ]

    return JsonResponse({'messages': mensajes_data})


@login_required
def seguimiento_adopcion(request, adopcion_id):
    """Seguimiento después de una adopción para validar viabilidad"""
    try:
        adopcion = Adopcion.objects.get(idAdopcion=adopcion_id, idPropietario=request.user.idUsuario)
    except Adopcion.DoesNotExist:
        messages.error(request, 'Adopción no encontrada')
        return redirect('mis_adopciones')

    if request.method == 'POST':
        reporte = request.POST.get('reporte', '').strip()
        if reporte:
            adopcion.Control_Adopción = f'{adopcion.Control_Adopción}\n[{datetime.now().strftime("%Y-%m-%d %H:%M")}] {reporte}'
            adopcion.save()
            messages.success(request, 'Seguimiento guardado correctamente')
            return redirect('seguimiento_adopcion', adopcion_id=adopcion_id)
        else:
            messages.error(request, 'El reporte no puede estar vacío')

    context = {'adopcion': adopcion}
    return render(request, 'adopcion/seguimiento_adopcion.html', context)


@login_required
def descargar_excel_adopciones(request):
    """Descargar un Excel con todas las adopciones para administrador"""
    if request.user.tipo != 'Administrador':
        messages.error(request, 'No tienes permiso para esta acción')
        return redirect('index')

    adopciones = Adopcion.objects.select_related('idMascota').all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Adopciones'

    headers = [
        'ID', 'Mascota', 'Propietario', 'Criador', 'Estado', 'Solicitud', 'Adopción', 'Entrega', 'Motivo', 'Estado de solicitud', 'Control', 'Salud', 'Vivienda', 'Fuente'
    ]
    ws.append(headers)

    for adopcion in adopciones:
        propietario = Usuario.objects.filter(idUsuario=adopcion.idPropietario).first()
        criador = Usuario.objects.filter(idUsuario=adopcion.idCriador).first()
        ws.append([
            adopcion.idAdopcion,
            adopcion.idMascota.Nombre_Mascota,
            propietario.nombre + ' ' + propietario.apellido if propietario else 'N/A',
            criador.nombre + ' ' + criador.apellido if criador else 'N/A',
            adopcion.Estado,
            str(adopcion.Fecha_Solicitud),
            str(adopcion.Fecha_Adopción),
            str(adopcion.Fecha_Entrega),
            adopcion.Motivo_Adopción,
            adopcion.Estado_Solicitud,
            adopcion.Control_Adopción,
            adopcion.Estado_Salud_Mascota,
            adopcion.Lugar_Vivienda,
            adopcion.Fuente_Mascota,
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=adopciones_admin.xlsx'
    wb.save(response)
    return response


@login_required
def crear_adoptar_mascotas_random(request):
    """Crea 100 mascotas aleatorias para llenar la app."""
    if not request.user.is_authenticated:
        messages.error(request, 'Debes estar autenticado para usar esta función')
        return redirect('index')

    nombres_base = ['Luna', 'Nala', 'Max', 'Simba', 'Milo', 'Bella', 'Chispa', 'Rocky', 'Lola', 'Thor',
        'Zeus', 'Kira', 'Oliver', 'Mia', 'Toby', 'Coco', 'Maya', 'Leo', 'Lucy', 'Daisy']
    adjetivos = ['Brisa', 'Luna', 'Sol', 'Dorado', 'Nube', 'Cielo', 'Feliz', 'Miel', 'Dulce', 'Amor']
    nombres = [f'{nombre} {adjetivo}' for nombre in nombres_base for adjetivo in adjetivos]
    random.shuffle(nombres)
    razas = ['Labrador', 'Poodle', 'Shih Tzu', 'Salchicha', 'Siames', 'Persa', 'Sphynx', 'Golden', 'Border Collie', 'Husky']
    especies = ['Perro', 'Gato']

    for i in range(100):
        Mascota.objects.create(
            Nombre_Mascota=nombres[i],
            Fecha_Nacimiento=date.today(),
            Raza=random.choice(razas),
            Genero=random.choice(['Macho', 'Hembra']),
            Peso=round(random.uniform(3, 40), 2),
            Especie=random.choice(especies),
            Color=random.choice(['Marrón', 'Negro', 'Blanco', 'Atigrado', 'Dorado']),
            Tamaño=random.choice(['Pequeño', 'Mediano', 'Grande']),
            Historial_Mascota='Mascota de prueba generada para evaluación.',
            Origen='Criador',
            Tipo_Alimentación='Comida balanceada',
            Enfermedades='Ninguna conocida',
            Vivienda='Casa con patio',
            Vacunas='Al día',
            Compatibilidad_Mascota='Buena con niños y otros animales',
            Descripción_Física='Pelaje suave y brillante',
            idCriador=request.user.idUsuario,
            Estado_Salud='Buena',
            Esterilizado=True,
            Socializado=True,
            disponible=True,
            puntuacion=round(random.uniform(3.0, 5.0), 2),
            numero_personas_interesadas=random.randint(0, 18)
        )

    messages.success(request, '100 mascotas de prueba generadas correctamente.')
    return redirect('mascotas_adopcion_disponibles')


@login_required
def descargar_reporte_adopcion(request, adopcion_id):
    """Vista para descargar reporte de adopción en PDF"""
    try:
        adopcion = Adopcion.objects.get(idAdopcion=adopcion_id)
    except Adopcion.DoesNotExist:
        messages.error(request, 'Adopción no encontrada')
        return redirect('mis_adopciones')

    if adopcion.Estado != 'Aprobada':
        messages.warning(request, 'El reporte solo está disponible para adopciones aprobadas')
        return redirect('mis_adopciones')

    return generar_pdf_adopcion(adopcion)


@login_required
def solicitudes_adopcion_criador(request):
    """Vista para que criadores vean solicitudes de adopción de sus mascotas"""
    from usuarios.models import Criador
    
    usuario = request.user
    try:
        criador = Criador.objects.get(user=usuario)
    except:
        messages.error(request, "Solo los criadores pueden acceder a esta función")
        return redirect('index')
    
    # Obtener todas las solicitudes de adopción para mascotas del criador
    solicitudes = Adopcion.objects.filter(
        idCriador=usuario.idUsuario,
        Estado='Pendiente'
    ).order_by('-Fecha_Solicitud')
    
    # Obtener todas las mascotas del criador
    mascotas_criador = Mascota.objects.filter(idCriador=usuario.idUsuario)
    
    # Obtener solicitudes aprobadas también
    solicitudes_aprobadas = Adopcion.objects.filter(
        idCriador=usuario.idUsuario,
        Estado='Aprobada'
    ).order_by('-Fecha_Solicitud')
    
    for solicitud in list(solicitudes) + list(solicitudes_aprobadas):
        actualizar_estado_entrega(solicitud)

    context = {
        'solicitudes_pendientes': solicitudes,
        'solicitudes_aprobadas': solicitudes_aprobadas,
        'total_pendientes': solicitudes.count(),
        'total_aprobadas': solicitudes_aprobadas.count(),
    }
    return render(request, 'adopcion/solicitudes_adopcion.html', context)


@login_required
def aprobar_solicitud_adopcion(request, adopcion_id):
    """Vista para que criadores aprueben solicitudes de adopción"""
    from usuarios.models import Criador, Usuario
    from .models import Notificacion
    
    usuario = request.user
    try:
        criador = Criador.objects.get(user=usuario)
    except:
        messages.error(request, "Solo los criadores pueden aprobar solicitudes")
        return redirect('index')
    
    try:
        adopcion = Adopcion.objects.get(idAdopcion=adopcion_id, idCriador=usuario.idUsuario, Estado='Pendiente')
        adopcion.Estado = 'Aprobada'
        adopcion.Estado_Solicitud = 'En camino'
        adopcion.Fecha_Adopción = date.today()
        adopcion.Fecha_Entrega = date.today() + timedelta(days=5)
        adopcion.idMascota.disponible = False
        adopcion.idMascota.save()
        adopcion.save()
        
        # Crear notificación para el propietario
        try:
            propietario = Usuario.objects.get(idUsuario=adopcion.idPropietario)
            Notificacion.objects.create(
                usuario=propietario,
                tipo='adopcion_aprobada',
                adopcion=adopcion,
                titulo=f'Tu solicitud de adopcion fue aprobada',
                mensaje=f'Tu solicitud para adoptar a {adopcion.idMascota.Nombre_Mascota} ha sido aprobada. La mascota sera entregada en los proximos 5 dias. Fecha de entrega: {adopcion.Fecha_Entrega.strftime("%d de %B de %Y")}.',
                enlace_accion=f'/adopcion/detalles/{adopcion_id}/'
            )
        except Usuario.DoesNotExist:
            pass
        
        messages.success(request, f"Adopcion de {adopcion.idMascota.Nombre_Mascota} aprobada. La mascota se entregara en los proximos dias.")
        return redirect('solicitudes_adopcion_criador')
    except Adopcion.DoesNotExist:
        messages.error(request, "Solicitud de adopción no encontrada o no es tuya")
        return redirect('solicitudes_adopcion_criador')


@login_required
def rechazar_solicitud_adopcion(request, adopcion_id):
    """Vista para que criadores rechacen solicitudes de adopción"""
    from usuarios.models import Criador, Usuario
    from .models import Notificacion
    
    usuario = request.user
    try:
        criador = Criador.objects.get(user=usuario)
    except:
        messages.error(request, "Solo los criadores pueden rechazar solicitudes")
        return redirect('index')
    
    if request.method != 'POST':
        messages.error(request, 'Debes enviar la razón del rechazo.')
        return redirect('solicitudes_adopcion_criador')

    motivo_rechazo = request.POST.get('motivo_rechazo', '').strip()
    if not motivo_rechazo:
        messages.error(request, 'El motivo del rechazo es obligatorio.')
        return redirect('solicitudes_adopcion_criador')

    try:
        adopcion = Adopcion.objects.get(idAdopcion=adopcion_id, idCriador=usuario.idUsuario, Estado='Pendiente')
        adopcion.Estado = 'Rechazada'
        adopcion.Estado_Solicitud = 'Cancelada'
        adopcion.Motivo_Rechazo = motivo_rechazo
        adopcion.save()
        
        # Crear notificación para el propietario
        try:
            propietario = Usuario.objects.get(idUsuario=adopcion.idPropietario)
            Notificacion.objects.create(
                usuario=propietario,
                tipo='adopcion_rechazada',
                adopcion=adopcion,
                titulo=f'Tu solicitud de adopcion fue rechazada',
                mensaje=f'Tu solicitud para adoptar a {adopcion.idMascota.Nombre_Mascota} ha sido rechazada. Motivo: {motivo_rechazo}',
                enlace_accion=f'/adopcion/detalles/{adopcion_id}/'
            )
        except Usuario.DoesNotExist:
            pass
        
        messages.success(request, f"Solicitud de adopcion rechazada. Se ha notificado al solicitante.")
        return redirect('solicitudes_adopcion_criador')
    except Adopcion.DoesNotExist:
        messages.error(request, "Solicitud de adopción no encontrada o no es tuya")
        return redirect('solicitudes_adopcion_criador')


# ==================== CALIFICACIONES ====================

def actualizar_promedio_usuario(usuario_id):
    """
    Función helper que recalcula el promedio de calificaciones de un usuario
    basado en todas las calificaciones que ha recibido.
    """
    from usuarios.models import Usuario
    from django.db.models import Avg
    
    try:
        usuario = Usuario.objects.get(idUsuario=usuario_id)
        calificaciones = Calificacion.objects.filter(usuario_calificado_id=usuario_id)
        
        if calificaciones.exists():
            promedio = calificaciones.aggregate(avg_puntuacion=Avg('puntuacion'))['avg_puntuacion']
            total = calificaciones.count()
            usuario.puntuacion_promedio = round(promedio, 2)
            usuario.total_calificaciones = total
        else:
            usuario.puntuacion_promedio = 0
            usuario.total_calificaciones = 0
        
        usuario.save()
    except Usuario.DoesNotExist:
        pass


def obtener_calificaciones_recibidas(usuario_id, limite=5):
    """
    Obtiene las últimas calificaciones recibidas por un usuario.
    Retorna: lista de Calificacion objs con join a usuarios
    """
    return Calificacion.objects.filter(
        usuario_calificado_id=usuario_id
    ).select_related(
        'usuario_califica', 'adopcion', 'adopcion__idMascota'
    ).order_by('-fecha_creacion')[:limite]


def obtener_calificaciones_dadas(usuario_id, limite=5):
    """
    Obtiene las últimas calificaciones dadas por un usuario.
    """
    return Calificacion.objects.filter(
        usuario_califica_id=usuario_id
    ).select_related(
        'usuario_calificado', 'adopcion', 'adopcion__idMascota'
    ).order_by('-fecha_creacion')[:limite]


@login_required
def crear_calificacion(request, adopcion_id):
    """
    Vista para que un usuario califique a otro usuario después de una adopción.
    GET: Muestra el formulario de calificación
    POST: Guarda la calificación y actualiza el promedio
    """
    try:
        adopcion = Adopcion.objects.get(idAdopcion=adopcion_id)
    except Adopcion.DoesNotExist:
        messages.error(request, "La adopción no existe")
        return redirect('mis_adopciones')
    
    # Determinar quién califica a quién
    usuario_actual = request.user
    
    # El propietario califica al criador y viceversa
    if usuario_actual.idUsuario == adopcion.idPropietario:
        usuario_a_calificar_id = adopcion.idCriador
        rol_usuario = "Propietario"
    elif usuario_actual.idUsuario == adopcion.idCriador:
        usuario_a_calificar_id = adopcion.idPropietario
        rol_usuario = "Criador"
    else:
        messages.error(request, "No tienes permiso para calificar esta adopción")
        return redirect('mis_adopciones')
    
    # Obtener el usuario a calificar
    from usuarios.models import Usuario
    try:
        usuario_a_calificar = Usuario.objects.get(idUsuario=usuario_a_calificar_id)
    except Usuario.DoesNotExist:
        messages.error(request, "El usuario a calificar no existe")
        return redirect('mis_adopciones')
    
    # Verificar si la calificación ya existe
    calificacion_existente = Calificacion.objects.filter(
        usuario_califica=usuario_actual,
        usuario_calificado=usuario_a_calificar,
        adopcion=adopcion
    ).first()
    
    if request.method == 'POST':
        form = CalificacionForm(request.POST)
        if form.is_valid():
            puntuacion = int(form.cleaned_data['puntuacion'])
            comentario = form.cleaned_data.get('comentario', '')
            
            if calificacion_existente:
                # No permitir modificación (según requisitos, es definitiva)
                messages.error(request, "Ya has calificado esta adopción y no puede ser modificada")
                return redirect('mis_adopciones')
            
            # Crear nueva calificación
            nueva_calificacion = Calificacion.objects.create(
                usuario_califica=usuario_actual,
                usuario_calificado=usuario_a_calificar,
                adopcion=adopcion,
                puntuacion=puntuacion,
                comentario=comentario if comentario else None
            )
            
            # Actualizar promedio del usuario calificado
            actualizar_promedio_usuario(usuario_a_calificar.idUsuario)
            
            # Crear notificación para el usuario calificado
            from .models import Notificacion
            Notificacion.objects.create(
                usuario=usuario_a_calificar,
                tipo='calificacion',
                titulo=f'⭐ Nueva calificación de {usuario_actual.nombre}',
                mensaje=f'{usuario_actual.nombre} te ha calificado con {puntuacion} estrella{"s" if puntuacion != 1 else ""} por la adopción de {adopcion.idMascota.Nombre_Mascota}',
                relacionado_con=str(nueva_calificacion.idCalificacion)
            )
            
            messages.success(request, f"Has calificado a {usuario_a_calificar.nombre} exitosamente")
            return redirect('mis_adopciones')
    else:
        form = CalificacionForm()
    
    context = {
        'form': form,
        'adopcion': adopcion,
        'usuario_a_calificar': usuario_a_calificar,
        'calificacion_existente': calificacion_existente,
        'rol_usuario': rol_usuario,
    }
    
    return render(request, 'adopcion/calificar.html', context)


# ==================== API VIEWSETS ====================

class StandardResultsSetPagination(pagination.PageNumberPagination):
    page_size = 12
    page_size_query_param = 'page_size'
    max_page_size = 1000


class MascotaViewSet(viewsets.ModelViewSet):
    queryset = Mascota.objects.all()
    serializer_class = MascotaSerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    filter_fields = ['Especie', 'Genero', 'Tamaño', 'Estado_Salud']
    search_fields = ['Nombre_Mascota', 'Raza', 'Especie']
    ordering_fields = ['Fecha_Nacimiento', 'Nombre_Mascota']
    ordering = ['-Fecha_Nacimiento']
    
    @action(detail=True, methods=['get'])
    def disponibles(self, request):
        """Obtener solo mascotas disponibles para adopción"""
        mascotas = Mascota.objects.exclude(adopcion__Estado='Aprobada').distinct()
        serializer = self.get_serializer(mascotas, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def solicitar_adopcion(self, request, pk=None):
        """Crear solicitud de adopción para una mascota"""
        mascota = self.get_object()
        from datetime import date
        
        try:
            adopcion = Adopcion.objects.create(
                idMascota=mascota,
                idPropietario=request.user.idUsuario,
                Estado='Pendiente',
                Fecha_Solicitud=date.today(),
                Fecha_Adopción=date.today(),
                Fecha_Entrega=date.today(),
                Estado_Solicitud='En revisión',
                Motivo_Adopción=request.data.get('motivo', ''),
                Fuente_Mascota='Criador',
                Control_Adopción='',
                Estado_Salud_Mascota='',
                Lugar_Vivienda='',
                Info_Mascota='',
                Estado_Ingreso_Mascota='',
                Devolución='',
            )
            return Response(
                {'mensaje': 'Solicitud de adopción creada', 'id': adopcion.idAdopcion},
                status=status.HTTP_201_CREATED
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


class AdopcionViewSet(viewsets.ModelViewSet):
    serializer_class = AdopcionSerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [permissions.IsAuthenticated]
    filter_fields = ['Estado', 'Estado_Solicitud', 'Fuente_Mascota']
    ordering_fields = ['Fecha_Solicitud']
    ordering = ['-Fecha_Solicitud']
    
    def get_queryset(self):
        """Filtrar adopciones del usuario actual"""
        user = self.request.user
        return Adopcion.objects.filter(idPropietario=user.idUsuario)
    
    @action(detail=False, methods=['get'])
    def mis_adopciones(self, request):
        """Obtener las adopciones del usuario actual"""
        adopciones = self.get_queryset()
        serializer = self.get_serializer(adopciones, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def aprobar(self, request, pk=None):
        """Aprobar una solicitud de adopción"""
        adopcion = self.get_object()
        adopcion.Estado = 'Aprobada'
        adopcion.Estado_Solicitud = 'Completada'
        adopcion.save()
        serializer = self.get_serializer(adopcion)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def rechazar(self, request, pk=None):
        """Rechazar una solicitud de adopción"""
        adopcion = self.get_object()
        adopcion.Estado = 'Rechazada'
        adopcion.Estado_Solicitud = 'Cancelada'
        adopcion.save()
        serializer = self.get_serializer(adopcion)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def descargar_pdf(self, request, pk=None):
        """Descargar reporte de adopción en PDF"""
        adopcion = self.get_object()
        return generar_pdf_adopcion(adopcion)


# ==================== FUNCIONES PARA GENERAR PDF ====================

def generar_pdf_adopcion(adopcion):
    """Genera un PDF con el reporte de adopción"""
    buffer = BytesIO()
    
    # Crear documento PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.75*inch,
        leftMargin=0.75*inch,
        topMargin=0.75*inch,
        bottomMargin=0.75*inch,
    )
    
    # Lista para almacenar elementos del PDF
    elementos = []
    
    # Estilos
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle(
        'titulo_custom',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#e83e8c'),
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitulo_style = ParagraphStyle(
        'subtitulo_custom',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#a8419f'),
        spaceAfter=12,
        alignment=TA_CENTER,
    )
    
    heading_style = ParagraphStyle(
        'heading_custom',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=10,
        fontName='Helvetica-Bold',
    )
    
    normal_style = ParagraphStyle(
        'normal_custom',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6,
    )
    
    # Logo y título
    logo_path = find('images/Hairylove.png') or find('Hairylove.png')
    if logo_path:
        try:
            logo = Image(logo_path, width=80, height=80)
            logo.hAlign = 'CENTER'
            elementos.append(logo)
        except:
            pass

    elementos.append(Paragraph('🐾 REPORTE DE ADOPCIÓN', titulo_style))
    elementos.append(Spacer(1, 0.1*inch))
    elementos.append(Paragraph('Hairy Love | Informe oficial de proceso de adopción', subtitulo_style))
    elementos.append(Spacer(1, 0.15*inch))
    
    # Fecha de generación
    fecha_generacion = datetime.now().strftime('%d de %B de %Y a las %H:%M:%S')
    elementos.append(Paragraph(f'Generado: {fecha_generacion}', styles['Normal']))
    elementos.append(Spacer(1, 0.25*inch))
    
    # Información de la adopción
    elementos.append(Paragraph('INFORMACIÓN DE LA ADOPCIÓN', heading_style))
    
    info_adopcion = [
        ['ID de Adopción:', str(adopcion.idAdopcion)],
        ['Estado:', adopcion.Estado],
        ['Estado de Solicitud:', adopcion.Estado_Solicitud],
        ['Fecha de Solicitud:', str(adopcion.Fecha_Solicitud)],
        ['Fecha de Adopción:', str(adopcion.Fecha_Adopción)],
        ['Fecha de Entrega:', str(adopcion.Fecha_Entrega)],
    ]
    
    tabla_adopcion = Table(info_adopcion, colWidths=[2.5*inch, 3.5*inch])
    tabla_adopcion.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5e6e1')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elementos.append(tabla_adopcion)
    elementos.append(Spacer(1, 0.25*inch))

    # Información del propietario y criador
    propietario = Usuario.objects.filter(idUsuario=adopcion.idPropietario).first()
    criador = Usuario.objects.filter(idUsuario=adopcion.idCriador).first()

    if propietario or criador:
        elementos.append(Paragraph('INFORMACIÓN DE USUARIOS', heading_style))
        usuarios_data = []
        if propietario:
            usuarios_data += [
                ['Propietario:', f'{propietario.nombre} {propietario.apellido}'],
                ['Correo propietario:', propietario.correo],
            ]
        if criador:
            usuarios_data += [
                ['Criador:', f'{criador.nombre} {criador.apellido}'],
                ['Correo criador:', criador.correo],
            ]

        tabla_usuarios = Table(usuarios_data, colWidths=[2.5*inch, 3.5*inch])
        tabla_usuarios.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5e6e1')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elementos.append(tabla_usuarios)
        elementos.append(Spacer(1, 0.25*inch))

    # Línea de tiempos
    elementos.append(Paragraph('CRONOGRAMA DEL PROCESO', heading_style))
    timeline_data = [
        ['Solicitud recibida', str(adopcion.Fecha_Solicitud)],
        ['Fecha de aprobación/confirmación', str(adopcion.Fecha_Adopción)],
        ['Fecha estimada de entrega', str(adopcion.Fecha_Entrega)],
        ['Estado del proceso', adopcion.Estado_Solicitud],
    ]
    tabla_timeline = Table(timeline_data, colWidths=[2.5*inch, 3.5*inch])
    tabla_timeline.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5e6e1')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elementos.append(tabla_timeline)
    elementos.append(Spacer(1, 0.25*inch))

    # Información de la mascota
    elementos.append(Paragraph('INFORMACIÓN DE LA MASCOTA', heading_style))
    
    try:
        mascota = adopcion.idMascota
        info_mascota = [
            ['Nombre:', mascota.Nombre_Mascota],
            ['Especie:', mascota.Especie],
            ['Raza:', mascota.Raza],
            ['Género:', mascota.Genero],
            ['Tamaño:', mascota.Tamaño],
            ['Estado de Salud:', mascota.Estado_Salud],
            ['Fecha de Nacimiento:', str(mascota.Fecha_Nacimiento)],
        ]
        
        tabla_mascota = Table(info_mascota, colWidths=[2.5*inch, 3.5*inch])
        tabla_mascota.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5e6e1')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elementos.append(tabla_mascota)
        elementos.append(Spacer(1, 0.25*inch))

        elementos.append(Paragraph('DETALLES ADICIONALES', heading_style))
        detalles_adicionales = [
            ['Origen de la mascota:', mascota.Origen or 'N/A'],
            ['Tipo de alimentación:', mascota.Tipo_Alimentación or 'N/A'],
            ['Condición de salud:', adopcion.Estado_Salud_Mascota or 'N/A'],
            ['Lugar de vivienda del solicitante:', adopcion.Lugar_Vivienda or 'N/A'],
            ['Resumen de acompañamiento:', adopcion.Info_Mascota or 'N/A'],
        ]
        tabla_detalles = Table(detalles_adicionales, colWidths=[2.5*inch, 3.5*inch])
        tabla_detalles.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5e6e1')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elementos.append(tabla_detalles)
    except:
        elementos.append(Paragraph('Información de mascota no disponible', styles['Normal']))
    
    elementos.append(Spacer(1, 0.25*inch))
    
    # Información adicional
    if adopcion.Motivo_Adopción:
        elementos.append(Paragraph('MOTIVO DE ADOPCIÓN', heading_style))
        elementos.append(Paragraph(adopcion.Motivo_Adopción, normal_style))
        elementos.append(Spacer(1, 0.15*inch))
    
    # Pie de página
    elementos.append(Spacer(1, 0.35*inch))
    elementos.append(Paragraph('Este documento es un reporte oficial de Hairy Love', styles['Normal']))
    elementos.append(Paragraph('Para más información, visite www.hairylove.com', styles['Normal']))
    
    # Construir PDF
    doc.build(elementos)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="adopcion_{adopcion.idAdopcion}_reporte.pdf"'
    
    return response


@login_required
def carga_masiva_mascotas(request):
    """Vista para carga masiva de mascotas desde Excel"""
    from usuarios.models import Criador
    
    usuario = request.user
    try:
        criador = Criador.objects.get(user=usuario)
    except Criador.DoesNotExist:
        messages.error(request, "Solo los criadores pueden cargar mascotas masivamente")
        return redirect('index')
    
    if request.method == 'POST':
        form = CargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo_excel']
            
            # Validar extensión del archivo
            if not archivo.name.endswith('.xlsx'):
                messages.error(request, "Solo se permiten archivos .xlsx")
                return redirect('carga_masiva_mascotas')

            try:    
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                
                # Leer encabezados de la primera fila
                headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                headers = [str(h).strip().lower() if h else '' for h in headers_row]
                
                # Crear diccionario de índices de columnas por nombre
                column_indices = {}
                for i, header in enumerate(headers):
                    if header:
                        # Mapear diferentes variaciones de nombres de columna
                        if 'nombre' in header and 'mascota' in header:
                            column_indices['nombre'] = i
                        elif 'fecha' in header and ('nac' in header or 'nacimiento' in header):
                            column_indices['fecha_nac'] = i
                        elif header == 'raza':
                            column_indices['raza'] = i
                        elif header == 'genero' or header == 'género':
                            column_indices['genero'] = i
                        elif header == 'peso':
                            column_indices['peso'] = i
                        elif header == 'especie':
                            column_indices['especie'] = i
                        elif header == 'color':
                            column_indices['color'] = i
                        elif 'tamaño' in header or 'tamano' in header or header == 'size':
                            column_indices['tamaño'] = i
                        elif 'historial' in header:
                            column_indices['historial'] = i
                        elif 'alimentación' in header or 'alimentacion' in header or 'tipo_alim' in header:
                            column_indices['tipo_alim'] = i
                        elif 'enfermedad' in header:
                            column_indices['enfermedades'] = i
                        elif header == 'vivienda':
                            column_indices['vivienda'] = i
                        elif header == 'vacunas':
                            column_indices['vacunas'] = i
                        elif 'compatibilidad' in header:
                            column_indices['compatibilidad'] = i
                        elif 'descripción' in header or 'descripcion' in header or 'desc_fisica' in header:
                            column_indices['desc_fisica'] = i
                        elif 'estado_salud' in header or 'estado salud' in header:
                            column_indices['estado_salud'] = i
                        elif header == 'origen':
                            column_indices['origen'] = i
                        elif 'esterilizado' in header:
                            column_indices['esterilizado'] = i
                        elif 'socializado' in header:
                            column_indices['socializado'] = i
                        elif 'url' in header or 'imagen' in header or 'image' in header:
                            column_indices['image_url'] = i
                
                print(f"Índices de columnas detectados: {column_indices}")

            except Exception as e:
                messages.error(request, f"Error al cargar el archivo Excel: {str(e)}")
                return redirect('carga_masiva_mascotas')

            errores = []
            mascotas_creadas = 0

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row_num <= 5:  # Debug first 5 rows
                    print(f"Row {row_num}: {row}")

                if not any(cell is not None and str(cell).strip() != '' for cell in row):
                    continue

                # Función auxiliar para obtener valor seguro de la fila
                def get_cell(col_name, default=None):
                    idx = column_indices.get(col_name)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        return val if val is not None else default
                    return default

                nombre = get_cell('nombre')
                fecha_nac_str = get_cell('fecha_nac')
                raza = get_cell('raza')
                genero = get_cell('genero')
                peso = get_cell('peso')
                especie = get_cell('especie')
                color = get_cell('color')
                tamaño = get_cell('tamaño')
                historial = get_cell('historial')
                tipo_alim = get_cell('tipo_alim')
                enfermedades = get_cell('enfermedades')
                vivienda = get_cell('vivienda')
                vacunas = get_cell('vacunas')
                compatibilidad = get_cell('compatibilidad')
                desc_fisica = get_cell('desc_fisica')
                estado_salud = get_cell('estado_salud')
                origen = get_cell('origen')
                esterilizado_val = get_cell('esterilizado')
                socializado_val = get_cell('socializado')
                image_url = get_cell('image_url')

                esterilizado = parse_boolean(esterilizado_val)
                socializado = parse_boolean(socializado_val)

                if not nombre or not raza or not genero:
                    errores.append(f"Fila {row_num}: Nombre, raza y género son obligatorios")
                    continue

                # Validar y mapear campos con choices
                genero = validate_and_map_genero(genero)
                tamaño = validate_and_map_tamaño(tamaño)
                estado_salud = validate_and_map_estado_salud(estado_salud)
                origen = validate_and_map_origen(origen)

                foto = None
                advertencia_imagen = None
                if image_url and str(image_url).strip():
                    try:
                        url = str(image_url).strip()
                        content, content_type = download_image_from_url(url)

                        ext = ''
                        if content_type and content_type.startswith('image/'):
                            ext = content_type.split('/')[-1].split(';')[0]
                        if not ext or '/' in ext or ext in ('html', 'htm', 'php', 'asp', 'aspx', 'jsp'):
                            ext = url.split('.')[-1].split('?')[0][:5].lower()
                        if not ext or '/' in ext or ext in ('html', 'htm', 'php', 'asp', 'aspx', 'jsp'):
                            ext = 'jpg'

                        filename = f"{uuid.uuid4()}.{ext}"
                        foto = ContentFile(content, name=filename)
                        print(f"✓ Imagen descargada para {nombre}: {filename}")
                    except HTTPError as e:
                        print(f"✗ HTTP Error {e.code} descargando imagen de {nombre}: {url}")
                        advertencia_imagen = f"No se pudo descargar la imagen (HTTP {e.code})"
                    except URLError as e:
                        print(f"✗ URL Error descargando imagen de {nombre}: {url} - {e.reason}")
                        advertencia_imagen = f"Error de conexión al descargar la imagen"
                    except Exception as e:
                        print(f"✗ Error desconocido para {nombre}: {str(e)}")
                        advertencia_imagen = f"Error con imagen - {str(e)}"

                fecha_nac = None
                if fecha_nac_str is not None and str(fecha_nac_str).strip() != '':
                    try:
                        fecha_nac = parse_excel_date(fecha_nac_str, wb)
                    except Exception:
                        errores.append(
                            f"Fila {row_num}: Fecha de nacimiento inválida (formatos válidos: YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY, MM/DD/YYYY, MM-DD-YYYY)"
                        )
                        continue

                # Truncar campos de texto largos para evitar errores de longitud
                historial = (str(historial) if historial else '')[:500]
                tipo_alim = (str(tipo_alim) if tipo_alim else '')[:100]
                enfermedades = (str(enfermedades) if enfermedades else '')[:500]
                vivienda = (str(vivienda) if vivienda else '')[:100]
                vacunas = (str(vacunas) if vacunas else '')[:500]
                compatibilidad = (str(compatibilidad) if compatibilidad else '')[:500]
                desc_fisica = (str(desc_fisica) if desc_fisica else '')[:500]

                mascota_data = {
                    'Nombre_Mascota': nombre,
                    'Fecha_Nacimiento': fecha_nac,
                    'Raza': raza,
                    'Genero': genero,
                    'Peso': float(peso) if peso else 0,
                    'Especie': especie or 'Perro',
                    'Color': color or '',
                    'Tamaño': tamaño,
                    'Historial_Mascota': historial,
                    'Tipo_Alimentación': tipo_alim,
                    'Enfermedades': enfermedades,
                    'Vivienda': vivienda,
                    'Vacunas': vacunas,
                    'Compatibilidad_Mascota': compatibilidad,
                    'Descripción_Física': desc_fisica,
                    'Estado_Salud': estado_salud,
                    'Origen': origen,
                    'Esterilizado': esterilizado or False,
                    'Socializado': socializado or True,
                    'idCriador': usuario.idUsuario,
                    'disponible': True
                }
                
                if foto:
                    mascota_data['foto_mascota'] = foto
                
                try:
                    Mascota.objects.create(**mascota_data)
                    mascotas_creadas += 1
                    # Si hubo advertencia con la imagen, agregarla a la lista pero no como error
                    if advertencia_imagen:
                        print(f"⚠️ Fila {row_num}: {nombre} creada, pero {advertencia_imagen}")
                except Exception as e:
                    errores.append(f"Fila {row_num}: No se pudo crear la mascota - {str(e)}")

            # Mostrar resultados después de procesar todas las filas
            if mascotas_creadas > 0:
                messages.success(request, f"✅ {mascotas_creadas} mascotas creadas correctamente.")
            if errores:
                mensajes_error = '; '.join(errores[:5])
                if len(errores) > 5:
                    mensajes_error += f"; y {len(errores) - 5} errores más"
                messages.warning(request, f"⚠️ {len(errores)} errores encontrados: {mensajes_error}")
            if mascotas_creadas == 0 and not errores:
                messages.info(request, "No se encontraron filas válidas para procesar.")
            return redirect('carga_masiva_mascotas')
        else:
            messages.error(request, "El formulario no es válido")
    
    # Si es GET, mostrar el formulario
    if request.method == 'GET':
        form = CargaMasivaForm()
    
    return render(request, 'adopcion/carga_masiva.html', {'form': form})

# ==================== APIS PARA REGISTRO DINÁMICO DE MASCOTAS ====================

def api_especies(request):
    """API que retorna todas las especies disponibles (Perro, Gato)"""
    return JsonResponse({
        'especies': ESPECIES
    })


def api_razas_por_especie(request):
    """API que retorna las razas según la especie seleccionada"""
    especie = request.GET.get('especie', '')
    
    if especie in RAZAS_POR_ESPECIE:
        razas = RAZAS_POR_ESPECIE[especie]
    else:
        razas = []
    
    return JsonResponse({
        'razas': razas
    })


def api_generos(request):
    """API que retorna los géneros disponibles"""
    generos = [
        {'value': 'Macho', 'label': 'Macho'},
        {'value': 'Hembra', 'label': 'Hembra'}
    ]
    return JsonResponse({
        'generos': generos
    })


def api_tamanos(request):
    """API que retorna los tamaños disponibles"""
    tamanos = [
        {'value': 'Pequeño', 'label': 'Pequeño'},
        {'value': 'Mediano', 'label': 'Mediano'},
        {'value': 'Grande', 'label': 'Grande'},
    ]
    return JsonResponse({
        'tamanos': tamanos
    })


# ==================== VISTAS PARA CRIADORES: EDITAR/ELIMINAR/SOLICITUDES ====================

@login_required
def editar_mascota(request, mascota_id):
    """Vista para que criadores editen sus mascotas"""
    try:
        mascota = Mascota.objects.get(idMascota=mascota_id)
    except Mascota.DoesNotExist:
        messages.error(request, "Mascota no encontrada")
        return redirect('mis_mascotas_adopcion')
    
    # Verificar que sea el propietario
    if mascota.idCriador != request.user.idUsuario:
        messages.error(request, "No tienes permiso para editar esta mascota")
        return redirect('detalles_mascota', mascota_id=mascota_id)
    
    if request.method == 'POST':
        form = MascotaAdopcionForm(request.POST, request.FILES, instance=mascota)
        print(f"Formulario válido: {form.is_valid()}")
        print(f"Datos POST: {request.POST}")
        print(f"Archivos: {request.FILES}")
        
        if form.is_valid():
            try:
                mascota = form.save(commit=False)
                mascota.fecha_actualizacion = timezone.now()
                mascota.save()
                print(f"Mascota guardada exitosamente: {mascota.Nombre_Mascota}")
                messages.success(request, f"Mascota {mascota.Nombre_Mascota} actualizada correctamente.")
                return redirect('detalles_mascota', mascota_id=mascota_id)
            except Exception as e:
                print(f"Error al guardar: {str(e)}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"Error al guardar los cambios: {str(e)}")
        else:
            print(f"Errores del formulario: {form.errors}")
            # Mostrar errores específicos del formulario
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            messages.error(request, f"Errores en el formulario: {'; '.join(error_messages)}")
    else:
        form = MascotaAdopcionForm(instance=mascota)
    
    context = {'form': form, 'mascota': mascota, 'es_edicion': True}
    return render(request, 'adopcion/editar_mascota.html', context)


@login_required
def eliminar_mascota(request, mascota_id):
    """Vista para eliminar una mascota"""
    try:
        mascota = Mascota.objects.get(idMascota=mascota_id)
    except Mascota.DoesNotExist:
        messages.error(request, "Mascota no encontrada")
        return redirect('mis_mascotas_adopcion')
    
    # Verificar que sea el propietario
    if mascota.idCriador != request.user.idUsuario:
        messages.error(request, "No tienes permiso para eliminar esta mascota")
        return redirect('detalles_mascota', mascota_id=mascota_id)
    
    if request.method == 'POST':
        nombre_mascota = mascota.Nombre_Mascota
        mascota.delete()
        messages.success(request, f"{nombre_mascota} ha sido eliminada.")
        return redirect('mis_mascotas_adopcion')
    
    # Mostrar confirmación
    context = {'mascota': mascota}
    return render(request, 'adopcion/confirmar_eliminar_mascota.html', context)


@login_required
def ver_solicitudes_mascota(request, mascota_id):
    """Vista para que criadores vean solicitudes de adopción de sus mascotas"""
    try:
        mascota = Mascota.objects.get(idMascota=mascota_id)
    except Mascota.DoesNotExist:
        messages.error(request, "Mascota no encontrada")
        return redirect('mis_mascotas_adopcion')
    
    # Verificar que sea el propietario
    if mascota.idCriador != request.user.idUsuario:
        messages.error(request, "No tienes permiso para ver estas solicitudes")
        return redirect('detalles_mascota', mascota_id=mascota_id)
    
    # Obtener todas las solicitudes de adopción para esta mascota
    solicitudes = Adopcion.objects.filter(
        idMascota=mascota
    ).order_by('-Fecha_Solicitud')
    
    # Obtener datos de propietarios
    solicitudes_con_usuario = []
    for solicitud in solicitudes:
        try:
            usuario = Usuario.objects.get(idUsuario=solicitud.idPropietario)
            solicitud.usuario_datos = usuario
        except Usuario.DoesNotExist:
            solicitud.usuario_datos = None
        solicitudes_con_usuario.append(solicitud)
    
    context = {
        'mascota': mascota,
        'solicitudes': solicitudes_con_usuario
    }
    return render(request, 'adopcion/solicitudes_mascota.html', context)
