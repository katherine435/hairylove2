import openpyxl

wb = openpyxl.load_workbook('mascotass.xlsx')
ws = wb.active
print('Headers:', [cell.value for cell in ws[1]])
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=5, values_only=True)):
    print(f'Row {i+2}:', row[:17])